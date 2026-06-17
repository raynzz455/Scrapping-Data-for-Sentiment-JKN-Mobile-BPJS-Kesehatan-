"""
Export helpers: generate CSV dan XLSX bytes untuk Streamlit download button.
"""
from __future__ import annotations
import io
from collections import Counter

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment


def to_csv_bytes(df: pd.DataFrame) -> bytes:
    buf = io.StringIO()
    df.to_csv(buf, index=False, encoding="utf-8-sig")
    return buf.getvalue().encode("utf-8-sig")


def to_xlsx_bytes(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        # ── Sheet 1: Data lengkap ──
        df.to_excel(writer, sheet_name="Ulasan Lengkap", index=False)
        ws = writer.sheets["Ulasan Lengkap"]
        _style_header(ws, "#111111", "#00E5A0")
        _color_rows(ws, df)
        _autofit(ws)

        # ── Sheet 2: Ringkasan ──
        if "sentiment" in df.columns:
            counts = Counter(df["sentiment"])
            total  = len(df)
            summary = pd.DataFrame({
                "Sentimen":    ["Positif 😊", "Netral 😐", "Negatif 😞", "Total"],
                "Jumlah":      [counts.get("positif", 0), counts.get("netral", 0),
                                counts.get("negatif", 0), total],
                "Persen (%)":  [
                    round(counts.get("positif", 0) / total * 100, 2) if total else 0,
                    round(counts.get("netral",  0) / total * 100, 2) if total else 0,
                    round(counts.get("negatif", 0) / total * 100, 2) if total else 0,
                    100.0,
                ],
            })
            summary.to_excel(writer, sheet_name="Ringkasan", index=False)
            _style_header(writer.sheets["Ringkasan"], "#111111", "#00E5A0")

        # ── Sheet 3: Tren Bulanan ──
        if "review_date" in df.columns and "sentiment" in df.columns:
            df2 = df.copy()
            df2["bulan"] = pd.to_datetime(df2["review_date"], errors="coerce").dt.to_period("M").astype(str)
            tren = df2.groupby(["bulan", "sentiment"]).size().unstack(fill_value=0).reset_index()
            tren.to_excel(writer, sheet_name="Tren Bulanan", index=False)
            _style_header(writer.sheets["Tren Bulanan"], "#111111", "#00E5A0")

    buf.seek(0)
    return buf.read()


# ── Internal helpers ──────────────────────────────────────

def _style_header(ws, bg_hex: str, fg_hex: str):
    bg = bg_hex.replace("#", "")
    fg = fg_hex.replace("#", "")
    for cell in ws[1]:
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.font      = Font(bold=True, color=fg, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _color_rows(ws, df: pd.DataFrame):
    fills = {
        "positif": PatternFill("solid", fgColor="0D2E1F"),
        "negatif": PatternFill("solid", fgColor="2E0D0D"),
        "netral":  PatternFill("solid", fgColor="2E2A00"),
    }
    if "sentiment" not in df.columns:
        return
    cols = list(df.columns)
    s_idx = cols.index("sentiment")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        val  = row[s_idx].value
        fill = fills.get(val)
        if fill:
            for cell in row:
                cell.fill = fill


def _autofit(ws, max_width: int = 55):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, max_width)
